##################################################################################
#                            Worldview Bulk Downloads

# Written by Colby Rand
# 2023/01/30

# This script downloads Maxar Worldview satellite imagery from NASA's Amazon Web
# Services (AWS) Simple Storage Service (S3). This script is an alternative to
# manually clicking on the provided URL inventory file with distribution links.

# Run the script by calling python from the terminal and then selecting the Excel
# file with the imagery that needs downloading:

# python bulk_downloads_worldview.py

##################################################################################

import os
import openpyxl
import subprocess
from tkinter import filedialog as fd
import tkinter as tk

# Function used to call wget
def runcmd(cmd, verbose = False, *args, **kwargs):

    process = subprocess.Popen(
        cmd,
        stdout = subprocess.PIPE,
        stderr = subprocess.PIPE,
        text = True,
        shell = True
    )
    std_out, std_err = process.communicate()
    if verbose:
        print(std_out.strip(), std_err)
    pass

# Get rid of tkinter root window
root = tk.Tk()
root.withdraw()

# Path to Excel file
filename = fd.askopenfilename()

# Load excel with its path
wrkbk = openpyxl.load_workbook(filename)
sh = wrkbk.active

path_name = filename.split('/')[4]
site_name = path_name.split('.')[0]

# Create a new folder to download the imagery to
os.chdir(os.path.expanduser('~')+'/Downloads/')
os.mkdir(site_name)
os.chdir(site_name)

# Iterate through the links in Excel and download the imagery
for i in range(1, sh.max_row+1):
    id = sh.cell(row=i, column=1)
    print('Downloading '+id.value.split('/')[5]+'...')
    runcmd('wget '+id.value, verbose = True)